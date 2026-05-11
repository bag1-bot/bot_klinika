from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

import openpyxl


def _s(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return str(v).strip() or None


def parse_price_xlsx(path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [ws.cell(1, c).value for c in range(1, 6)]

    section: str | None = None
    data: dict[str, list[dict[str, Any]]] = {}
    current: dict[str, Any] | None = None

    for r in range(2, ws.max_row + 1):
        a = _s(ws.cell(r, 1).value)  # Раздел
        b = _s(ws.cell(r, 2).value)  # Название услуги по прайсу
        c = _s(ws.cell(r, 3).value)  # Состоит из услуг
        d = _s(ws.cell(r, 4).value)  # Номенклатура
        e = ws.cell(r, 5).value  # Цена

        if a:
            section = a
            data.setdefault(section, [])
            current = None

        if not section:
            continue

        if b:
            if isinstance(e, (int, float)):
                price: Any = float(e)
            elif e is None:
                price = None
            else:
                price = str(e).strip() or None

            current = {
                "услуга": b,
                "цена": price,
                "состоит_из": [],
                "номенклатура": [],
            }
            if c:
                current["состоит_из"].append(c)
            if d:
                current["номенклатура"].append(d)
            data[section].append(current)
            continue

        if current is None:
            continue
        if c:
            current["состоит_из"].append(c)
        if d:
            current["номенклатура"].append(d)

    return {"headers": headers, "разделы": list(data.keys()), "данные": data}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--in",
        dest="in_path",
        default=str(Path("data/rag_docs/price.xlsx")),
        help="Path to price.xlsx",
    )
    ap.add_argument(
        "--out",
        dest="out_path",
        default=str(Path("data/rag_docs/price_parsed_full.json")),
        help="Output JSON path (compact)",
    )
    ap.add_argument(
        "--pretty-out",
        dest="pretty_out_path",
        default=str(Path("data/rag_docs/price_parsed_full.pretty.json")),
        help="Output JSON path (pretty)",
    )
    args = ap.parse_args()

    in_path = Path(args.in_path)
    out_path = Path(args.out_path)
    pretty_out_path = Path(args.pretty_out_path)

    payload = parse_price_xlsx(in_path)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    pretty_out_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"Wrote: {out_path}")
    print(f"Wrote: {pretty_out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

