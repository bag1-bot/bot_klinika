from __future__ import annotations

import json
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _is_empty(v: Any) -> bool:
    return v is None or v == ""


@dataclass(frozen=True)
class PriceRow:
    category: str
    service_name: str
    service: str
    includes: str
    price: float | int | str | None
    term: str | None


def _as_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def load_rows(path: Path, sheet_name: str | None = None) -> list[PriceRow]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    # Header is on row 1 in this file.
    start_row = 2
    max_r = ws.max_row

    current_category = ""
    current_service_name = ""

    rows: list[PriceRow] = []
    for r in range(start_row, max_r + 1):
        a = ws.cell(r, 1).value  # Категория
        b = ws.cell(r, 2).value  # Наименование Услуги
        c = ws.cell(r, 3).value  # Услуга
        d = ws.cell(r, 4).value  # Что входит в услугу
        e = ws.cell(r, 5).value  # Цена
        f = ws.cell(r, 6).value  # Срок оказания услуги

        if not _is_empty(a):
            current_category = _as_str(a)
        if not _is_empty(b):
            current_service_name = _as_str(b)

        service = _as_str(c)
        includes = _as_str(d)
        if _is_empty(c) and _is_empty(d) and _is_empty(e) and _is_empty(f):
            continue

        rows.append(
            PriceRow(
                category=current_category,
                service_name=current_service_name,
                service=service,
                includes=includes,
                price=e,
                term=_as_str(f) if not _is_empty(f) else None,
            )
        )

    # Drop trailing empty “blocks” if any.
    return [r for r in rows if r.category or r.service_name or r.service or r.includes]


def build_tree(rows: list[PriceRow]) -> dict[str, dict[str, list[PriceRow]]]:
    tree: dict[str, dict[str, list[PriceRow]]] = defaultdict(lambda: defaultdict(list))
    for r in rows:
        tree[r.category][r.service_name].append(r)
    return tree


def main() -> None:
    path = Path("data/rag_docs/price.xlsx")
    rows = load_rows(path)
    tree = build_tree(rows)

    # Compact summary for chat consumption.
    summary = {
        "file": str(path),
        "total_rows": len(rows),
        "categories": sorted([k for k in tree.keys() if k]),
        "counts": {
            "categories": len([k for k in tree.keys() if k]),
            "service_names_total": sum(len(v) for v in tree.values()),
        },
        "examples": [],
    }

    # Add a few example branches (first 3 categories, first 2 service names each).
    for cat in summary["categories"][:3]:
        service_names = sorted([k for k in tree[cat].keys() if k])
        for sn in service_names[:2]:
            branch_rows = tree[cat][sn][:8]
            summary["examples"].append(
                {
                    "category": cat,
                    "service_name": sn,
                    "rows_preview": [
                        {
                            "service": r.service,
                            "includes": r.includes,
                            "price": r.price,
                            "term": r.term,
                        }
                        for r in branch_rows
                    ],
                }
            )

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

