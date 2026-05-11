from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


@dataclass(frozen=True)
class SheetPreview:
    name: str
    max_row: int
    max_col: int
    first_non_empty_row: int | None
    rows: list[list[Any]]


def _is_empty(v: Any) -> bool:
    return v is None or v == ""


def _trim_trailing_empty(row: list[Any]) -> list[Any]:
    i = len(row)
    while i > 0 and _is_empty(row[i - 1]):
        i -= 1
    return row[:i]


def _first_non_empty_row(ws, search_rows: int) -> int | None:
    max_r = min(ws.max_row, search_rows)
    for r in range(1, max_r + 1):
        for c in range(1, ws.max_column + 1):
            if not _is_empty(ws.cell(r, c).value):
                return r
    return None


def preview_workbook(
    path: Path,
    *,
    data_only: bool = True,
    header_search_rows: int = 60,
    preview_rows: int = 30,
    preview_cols: int = 25,
) -> list[SheetPreview]:
    wb = load_workbook(path, data_only=data_only)
    previews: list[SheetPreview] = []

    for name in wb.sheetnames:
        ws = wb[name]
        first = _first_non_empty_row(ws, header_search_rows)
        start = first or 1
        end = min(ws.max_row, start + preview_rows - 1)

        rows: list[list[Any]] = []
        for r in range(start, end + 1):
            row = [ws.cell(r, c).value for c in range(1, min(ws.max_column, preview_cols) + 1)]
            row = _trim_trailing_empty(row)
            if not row:
                continue
            rows.append(row)

        previews.append(
            SheetPreview(
                name=name,
                max_row=ws.max_row,
                max_col=ws.max_column,
                first_non_empty_row=first,
                rows=rows,
            )
        )

    return previews


def main() -> None:
    xlsx_path = Path("data/rag_docs/price.xlsx")
    previews = preview_workbook(xlsx_path)

    print("file:", str(xlsx_path))
    print("sheets:", [p.name for p in previews])

    for p in previews:
        print()
        print("===", p.name, "===")
        print("max_row:", p.max_row, "max_col:", p.max_col)
        print("first_non_empty_row:", p.first_non_empty_row)
        for i, row in enumerate(p.rows, start=1):
            print(f"row#{i}:", row)


if __name__ == "__main__":
    main()

